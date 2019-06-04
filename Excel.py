from openpyxl import Workbook, load_workbook
from datetime import date

class Spreadsheet:
	def __init__(self, holdings, margin_initial_ratio, maintenance_ratio, buy_orders, sell_orders):
		#Sets up data
		#
		self.holdings = holdings
		self.margin_initial_ratio = margin_initial_ratio
		self.maintenance_ratio = maintenance_ratio
		self.buy_orders = buy_orders
		self.sell_orders = sell_orders
		self.date = date.today()
		self.wb = None
		self.ws = None
		self.row = None
		self.holdings_change = 0
		self.price_percent_change = 0
		self.file_name = "Tesla Stock Logs.xlsx"

	def new_workbook(self):
		#Sets up new workbook to record data for first time
		#
		self.wb = Workbook()
		self.ws = self.wb.active
		self.ws.column_dimensions["A"].width = 11
		self.ws['A1'] = "Date"
		self.ws['B1'] = "Robinhood Holdings"
		self.ws['C1'] = "Holdings Change"
		self.ws['D1'] = "Margin Initial Ratio"
		self.ws['E1'] = "Maintenance Ratio"
		self.ws['F1'] = "Buy Orders"
		self.ws['G1'] = "Sell Orders"
		self.ws['H1'] = "Net Orders"
		self.ws['I1'] = "Prices"
		self.ws['J1'] = "Price % Change"

	def load_workbook(self):
		#Opens up spreadsheet to write in new data
		#
		self.wb = load_workbook(self.file_name)
		self.ws = self.wb.active

	def calculate_data(self):
		self.row = self.ws.max_row + 1
		self.holdings_change = self.holdings - self.ws.cell(column=2, row=self.row-1).value
		self.price_percent_change = '=ROUND((I'+str(self.row)+'-I'+str(self.row-1)+')/I'+str(self.row-1)+'*100,1)&"%"'

	def save_data(self):
		#Saves data in spreadsheet
		#
		self.row = self.ws.max_row + 1
		self.ws.cell(column=1, row = self.row, value = self.date)
		self.ws.cell(column=2, row = self.row, value = self.holdings)
		self.ws.cell(column=3, row = self.row, value = self.holdings_change)
		self.ws.cell(column=4, row = self.row, value = self.margin_initial_ratio)
		self.ws.cell(column=5, row = self.row, value = self.maintenance_ratio)
		self.ws.cell(column=6, row = self.row, value = self.buy_orders)
		self.ws.cell(column=7, row = self.row, value = self.sell_orders)
		self.ws.cell(column=8, row = self.row, value = str(int("".join(letter for letter in self.buy_orders if letter not in ',')) - int("".join(letter for letter in self.sell_orders if letter not in ','))))
		self.ws.cell(column=9, row = self.row, value = '=FDS("TSLA","FG_PRICE("&A'+str(self.row)+'&","&A'+str(self.row)+'&")")') 
		self.ws.cell(column=10, row = self.row, value = self.price_percent_change)
		self.wb.save(self.file_name)